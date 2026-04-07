"""
Generate Excel test documents for E-EVKIN Modern production testing.
Uses real data from Rekap_Ver3.xlsx for Bojonggede and Cibinong puskesmas.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# REAL DATA FROM EXCEL
# ============================================================

BOJONGGEDE_DATA = [
    {"kode": "1.02.01.2.10.0001", "nama": "Pelayanan dan Penunjang Pelayanan BLUD", "sumber": "BLUD Puskesmas", "pagu": 3_738_528_000},
    {"kode": "1.02.02.2.02.0001", "nama": "Pengelolaan Pelayanan Kesehatan Ibu Hamil", "sumber": "DAK Non Fisik (BOK)", "pagu": 27_260_000},
    {"kode": "1.02.02.2.02.0005", "nama": "Pengelolaan Pelayanan Kesehatan pada Usia Pendidikan Dasar", "sumber": "DAK Non Fisik (BOK)", "pagu": 9_500_000},
    {"kode": "1.02.02.2.02.0010", "nama": "Pengelolaan Pelayanan Kesehatan Orang dengan Gangguan Jiwa Berat", "sumber": "DAK Non Fisik (BOK)", "pagu": 600_000},
    {"kode": "1.02.02.2.02.0011", "nama": "Pengelolaan Pelayanan Kesehatan Orang Terduga Tuberkulosis", "sumber": "DAK Non Fisik (BOK)", "pagu": 10_000_000},
    {"kode": "1.02.02.2.02.0015", "nama": "Pengelolaan Pelayanan Kesehatan Gizi Masyarakat", "sumber": "DAK Non Fisik (BOK)", "pagu": 97_130_000},
    {"kode": "1.02.02.2.02.0017", "nama": "Pengelolaan Pelayanan Kesehatan Lingkungan", "sumber": "DAK Non Fisik (BOK)", "pagu": 3_600_000},
    {"kode": "1.02.02.2.02.0020", "nama": "Pengelolaan Surveilans Kesehatan", "sumber": "DAK Non Fisik (BOK)", "pagu": 31_458_000},
    {"kode": "1.02.02.2.02.0021", "nama": "Pengelolaan Pelayanan Kesehatan Orang dengan Masalah Kesehatan Akibat Dampak Rokok", "sumber": "DAK Non Fisik (BOK)", "pagu": 2_000_000},
    {"kode": "1.02.02.2.02.0025", "nama": "Pelayanan Kesehatan Penyakit Menular dan Tidak Menular", "sumber": "DAK Non Fisik (BOK)", "pagu": 54_000_000},
    {"kode": "1.02.02.2.02.0033", "nama": "Operasional Pelayanan Puskesmas", "sumber": "DAK Non Fisik (BOK)", "pagu": 85_094_000},
    {"kode": "1.02.02.2.02.0033", "nama": "Operasional Pelayanan Puskesmas", "sumber": "APBD (PAD)", "pagu": 182_537_106},
    {"kode": "1.02.02.2.02.0046", "nama": "Pengelolaan Upaya Kesehatan Ibu dan Anak", "sumber": "DAK Non Fisik (BOK)", "pagu": 900_000},
    {"kode": "1.02.02.2.02.0048", "nama": "Pengelolaan Layanan Imunisasi", "sumber": "DAK Non Fisik (BOK)", "pagu": 43_700_000},
    {"kode": "1.02.05.2.01.0001", "nama": "Peningkatan Upaya Promosi Kesehatan, Advokasi, Kemitraan", "sumber": "DAK Non Fisik (BOK)", "pagu": 49_168_000},
    {"kode": "1.02.05.2.03.0001", "nama": "Bimbingan Teknis dan Supervisi Pengembangan UKBM", "sumber": "DAK Non Fisik (BOK)", "pagu": 32_364_000},
]

CIBINONG_DATA = [
    {"kode": "1.02.01.2.10.0001", "nama": "Pelayanan dan Penunjang Pelayanan BLUD", "sumber": "BLUD Puskesmas", "pagu": 2_135_455_000},
    {"kode": "1.02.02.2.02.0001", "nama": "Pengelolaan Pelayanan Kesehatan Ibu Hamil", "sumber": "DAK Non Fisik (BOK)", "pagu": 19_552_000},
    {"kode": "1.02.02.2.02.0005", "nama": "Pengelolaan Pelayanan Kesehatan pada Usia Pendidikan Dasar", "sumber": "DAK Non Fisik (BOK)", "pagu": 3_775_000},
    {"kode": "1.02.02.2.02.0010", "nama": "Pengelolaan Pelayanan Kesehatan Orang dengan Gangguan Jiwa Berat", "sumber": "DAK Non Fisik (BOK)", "pagu": 3_750_000},
    {"kode": "1.02.02.2.02.0011", "nama": "Pengelolaan Pelayanan Kesehatan Orang Terduga Tuberkulosis", "sumber": "DAK Non Fisik (BOK)", "pagu": 6_500_000},
    {"kode": "1.02.02.2.02.0015", "nama": "Pengelolaan Pelayanan Kesehatan Gizi Masyarakat", "sumber": "DAK Non Fisik (BOK)", "pagu": 94_490_000},
    {"kode": "1.02.02.2.02.0017", "nama": "Pengelolaan Pelayanan Kesehatan Lingkungan", "sumber": "DAK Non Fisik (BOK)", "pagu": 1_500_000},
    {"kode": "1.02.02.2.02.0020", "nama": "Pengelolaan Surveilans Kesehatan", "sumber": "DAK Non Fisik (BOK)", "pagu": 2_080_000},
    {"kode": "1.02.02.2.02.0021", "nama": "Pengelolaan Pelayanan Kesehatan Orang dengan Masalah Kesehatan Akibat Dampak Rokok", "sumber": "DAK Non Fisik (BOK)", "pagu": 8_416_000},
    {"kode": "1.02.02.2.02.0025", "nama": "Pelayanan Kesehatan Penyakit Menular dan Tidak Menular", "sumber": "DAK Non Fisik (BOK)", "pagu": 49_505_000},
    {"kode": "1.02.02.2.02.0033", "nama": "Operasional Pelayanan Puskesmas", "sumber": "DAK Non Fisik (BOK)", "pagu": 77_276_000},
    {"kode": "1.02.02.2.02.0033", "nama": "Operasional Pelayanan Puskesmas", "sumber": "APBD (PAD)", "pagu": 181_390_962},
    {"kode": "1.02.02.2.02.0046", "nama": "Pengelolaan Upaya Kesehatan Ibu dan Anak", "sumber": "DAK Non Fisik (BOK)", "pagu": 75_000},
    {"kode": "1.02.02.2.02.0048", "nama": "Pengelolaan Layanan Imunisasi", "sumber": "DAK Non Fisik (BOK)", "pagu": 27_952_500},
    {"kode": "1.02.05.2.01.0001", "nama": "Peningkatan Upaya Promosi Kesehatan, Advokasi, Kemitraan", "sumber": "DAK Non Fisik (BOK)", "pagu": 7_246_500},
    {"kode": "1.02.05.2.03.0001", "nama": "Bimbingan Teknis dan Supervisi Pengembangan UKBM", "sumber": "DAK Non Fisik (BOK)", "pagu": 22_976_000},
]

TOTAL_BJG = sum(d["pagu"] for d in BOJONGGEDE_DATA)
TOTAL_CBN = sum(d["pagu"] for d in CIBINONG_DATA)

# ============================================================
# STYLE HELPERS
# ============================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

def fmt_rp(val):
    """Format as Indonesian Rupiah."""
    return f"Rp {val:,.0f}".replace(",", ".")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def write_section_row(ws, row, text, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="top")
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL

def write_data_row(ws, row, values):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        cell.font = Font(name="Calibri", size=10)

def write_title(ws, row, text, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")

def write_subtitle(ws, row, text, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=11)
    cell.alignment = Alignment(horizontal="center")


# ============================================================
# PUSKESMAS TEST DOC GENERATOR
# ============================================================

TC_HEADERS = ["No", "ID", "Modul", "Nama Test Case", "Prasyarat", "Langkah Pengujian", "Contoh Input", "Output yang Diharapkan", "Status", "Catatan"]
TC_WIDTHS = [5, 14, 18, 30, 25, 45, 35, 50, 10, 25]


def generate_puskesmas_doc(filename, nama, username, kecamatan, wilayah, data, total_pagu, tc_prefix):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Test Cases ----
    ws = wb.active
    ws.title = "Test Cases"
    set_col_widths(ws, TC_WIDTHS)

    r = 1
    write_title(ws, r, f"DOKUMEN PENGUJIAN - PUSKESMAS {nama.upper()}", len(TC_HEADERS)); r += 1
    write_subtitle(ws, r, f"Aplikasi: E-EVKIN Modern | Role: Puskesmas | Username: {username} | Password: bogorkab", len(TC_HEADERS)); r += 1
    write_subtitle(ws, r, f"Kecamatan: {kecamatan} | Wilayah: {wilayah} | Tanggal: 30 Maret 2026 | Tahun Anggaran: 2026", len(TC_HEADERS)); r += 1
    r += 1

    write_header_row(ws, r, TC_HEADERS); r += 1

    # Top 3 sub-kegiatan for examples
    top3 = data[:3]
    top3_names = [d["nama"][:45] for d in top3]
    top3_sumber = [d["sumber"] for d in top3]
    top3_pagu = [d["pagu"] for d in top3]

    tc_num = 0
    def tc(modul, nama_tc, prasyarat, langkah, contoh, output):
        nonlocal r, tc_num
        tc_num += 1
        write_data_row(ws, r, [tc_num, f"{tc_prefix}-{tc_num:03d}", modul, nama_tc, prasyarat, langkah, contoh, output, "", ""])
        r += 1

    # --- LOGIN ---
    write_section_row(ws, r, "1. LOGIN & AUTENTIKASI", len(TC_HEADERS)); r += 1

    tc("Login", "Login Berhasil",
       "Aplikasi dapat diakses, belum login",
       "1. Buka /login\n2. Input username: {0}\n3. Input password: bogorkab\n4. Klik Login".format(username),
       f"Username: {username}\nPassword: bogorkab",
       f'Redirect ke /puskesmas/dashboard\nHeader: "Puskesmas {nama}"\nSidebar: Dashboard, Laporan, Target, Cara Pengisian')

    tc("Login", "Login Gagal - Password Salah",
       "Belum login",
       f"1. Input username: {username}\n2. Input password: wrongpass\n3. Klik Login",
       f"Username: {username}\nPassword: wrongpass",
       'Tetap di /login\nNotifikasi: "Username atau password salah"')

    tc("Login", "Akses Halaman Admin",
       f"Login sebagai {nama}",
       "1. Akses langsung URL /dashboard",
       "URL: /dashboard",
       "Redirect ke /puskesmas/dashboard atau 403")

    tc("Login", "Verifikasi Token (Refresh)",
       f"Sudah login sebagai {nama}",
       "1. Tekan F5 (refresh browser)",
       "(otomatis)",
       f"Tetap login sebagai {nama}\nDashboard tampil kembali")

    # --- DASHBOARD ---
    write_section_row(ws, r, "2. DASHBOARD PUSKESMAS", len(TC_HEADERS)); r += 1

    tc("Dashboard", "Tampilan Dashboard",
       f"Login sebagai {nama}",
       '1. Klik menu "Dashboard" di sidebar',
       "Klik sidebar > Dashboard",
       f"Filter: Tahun, Bulan\nStatistik Laporan Sendiri:\n- Total Laporan\n- Tersimpan (Draft)\n- Terkirim\n- Terverifikasi\nGrafik Budget: Target vs Realisasi\nTotal Target: {fmt_rp(total_pagu)}")

    tc("Dashboard", "Filter Dashboard per Tahun & Bulan",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026\n2. Pilih Bulan: Januari",
       "Tahun: 2026, Bulan: Januari",
       "Data di-refresh sesuai filter\nStatistik hanya Januari 2026")

    tc("Dashboard", "Grafik Budget YTD",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026\n2. Lihat grafik YTD",
       "Tahun: 2026",
       f"Grafik akumulasi bulanan\nTotal Target: {fmt_rp(total_pagu)}\nData spesifik {nama}")

    tc("Dashboard", "Budget Monthly Breakdown",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026, Bulan: Januari\n2. Lihat breakdown per sub kegiatan",
       "Tahun: 2026, Bulan: Januari",
       "Tabel breakdown:\n" + "\n".join([f"- {d['nama'][:40]} | {d['sumber']} | Target: {fmt_rp(d['pagu'])}" for d in data[:5]]) + "\n...")

    # --- MANAJEMEN TARGET ---
    write_section_row(ws, r, "3. MANAJEMEN TARGET", len(TC_HEADERS)); r += 1

    target_list = "\n".join([f"- {d['nama'][:40]} | {d['sumber']} | Pagu: {fmt_rp(d['pagu'])}" for d in data[:6]])
    tc("Target", "Lihat Target yang Di-assign",
       f"Login sebagai {nama}",
       '1. Klik menu "Target" di sidebar',
       "Klik sidebar > Target",
       f"Tabel target ({len(data)} baris):\n{target_list}\n... (total {len(data)} sub kegiatan x sumber)")

    tc("Target", "Input Target Baru",
       "Izin edit target_kinerja dan target_rp aktif",
       '1. Klik "Tambah Target"\n2. Isi data\n3. Simpan',
       f"Sub Kegiatan: {top3_names[1]}\nSumber: {top3_sumber[1]}\nSatuan: Orang\nTarget K: 240\nTarget Rp: {top3_pagu[1]:,.0f}\nTahun: 2026",
       '"Target berhasil disimpan"\nBaris baru muncul di tabel')

    tc("Target", "Edit Target Kinerja",
       "Izin edit target_kinerja aktif",
       f'1. Edit target "{top3_names[0][:30]}"\n2. Ubah Target K\n3. Simpan',
       "Target K diubah dari 12 menjadi 14",
       f'"Target berhasil diperbarui"\nTarget K: 14\nTarget Rp tetap: {fmt_rp(top3_pagu[0])}')

    tc("Target", "Bulk Input Target",
       "Izin edit target aktif",
       '1. Isi beberapa target sekaligus\n2. Klik "Simpan Semua"',
       "\n".join([f"Baris {i+1} ({d['nama'][:25]}): Target K = {12 if i==0 else 240}, Target Rp = {d['pagu']:,.0f}" for i, d in enumerate(top3)]),
       '"3 target berhasil disimpan"\nSemua baris ter-update')

    tc("Target", "Lihat Riwayat Target",
       "Target sudah pernah diubah",
       f'1. Klik icon riwayat pada "{top3_names[0][:30]}"',
       "Klik icon riwayat",
       f"Riwayat perubahan:\n- Versi 2 (30/03/2026): Target K: 14, oleh: {username}\n- Versi 1 (01/01/2026): Target K: 12, oleh: {username}")

    tc("Target", "Input Target Saat Izin Tidak Aktif",
       "Izin edit target_kinerja TIDAK aktif",
       "1. Coba ubah target\n2. Simpan",
       "Target K diubah menjadi 20",
       'Gagal menyimpan\n"Anda tidak diizinkan untuk mengedit target saat ini"')

    # --- ANGKAS ---
    write_section_row(ws, r, "4. MANAJEMEN ANGKAS", len(TC_HEADERS)); r += 1

    tc("Angkas", "Lihat Angkas",
       f"Login sebagai {nama}, di halaman Target",
       "1. Lihat bagian Angkas",
       "(otomatis tampil)",
       "Tabel angkas per sub kegiatan:\n" + "\n".join([f"- {d['nama'][:40]} | Angkas: {fmt_rp(int(d['pagu']*0.9))}" for d in data[:4]]))

    tc("Angkas", "Input Angkas",
       "Izin edit angkas aktif",
       "1. Edit angkas untuk sub kegiatan\n2. Simpan",
       f"Sub Kegiatan: {top3_names[0][:35]}\nAngkas: {int(top3_pagu[0]*0.95):,.0f}",
       f'"Angkas berhasil disimpan"\nAngkas: {fmt_rp(int(top3_pagu[0]*0.95))}')

    tc("Angkas", "Bulk Input Angkas",
       "Izin edit angkas aktif",
       "1. Isi beberapa angkas sekaligus\n2. Simpan",
       "\n".join([f"Baris {i+1}: Angkas = {int(d['pagu']*0.9):,.0f}" for i, d in enumerate(top3)]),
       '"3 angkas berhasil disimpan"')

    # --- INPUT LAPORAN ---
    write_section_row(ws, r, "5. INPUT LAPORAN (BULK)", len(TC_HEADERS)); r += 1

    tc("Laporan", "Buka Halaman Laporan Bulk",
       f"Login sebagai {nama}, izin edit aktif",
       '1. Klik menu "Laporan" di sidebar',
       "Klik sidebar > Laporan",
       f"Tabel bulk input:\n- Filter: Tahun, Bulan\n- {len(data)} baris sub kegiatan\n- Kolom: Sub Kegiatan, Sumber, Satuan, Target K, Target Rp, Angkas, Realisasi K, Realisasi Rp, Realisasi Fisik (%), Permasalahan, Upaya")

    tc("Laporan", "Input Laporan Satu Baris",
       "Di halaman Laporan, Tahun: 2026, Bulan: Januari",
       f'1. Pilih Tahun: 2026, Bulan: Januari\n2. Isi realisasi pada "{top3_names[0][:30]}"\n3. Klik Simpan',
       f"Baris: {top3_names[0][:35]} / {top3_sumber[0]}\nRealisasi K: 1\nRealisasi Rp: {int(top3_pagu[0]*0.08):,.0f}\nRealisasi Fisik: 8\nPermasalahan: Realisasi belanja masih rendah\nUpaya: Mempercepat proses pengadaan",
       f'"Laporan berhasil disimpan"\nStatus: Tersimpan (draft)\nTarget Rp: {fmt_rp(top3_pagu[0])}\nRealisasi Rp: {fmt_rp(int(top3_pagu[0]*0.08))}')

    tc("Laporan", "Bulk Input - Semua Baris",
       "Di halaman Laporan, Tahun: 2026, Bulan: Januari",
       '1. Isi semua baris\n2. Klik "Simpan Semua"',
       "\n".join([f"Baris {i+1}: {d['nama'][:30]} / {d['sumber']}\nRealisasi K: {1 if i==0 else 20+i*10}, Realisasi Rp: {int(d['pagu']*0.08):,.0f}, Fisik: {8+i*2}" for i, d in enumerate(data[:4])]),
       f'"{min(4, len(data))} laporan berhasil disimpan"\nSemua baris berstatus "Tersimpan"')

    tc("Laporan", "Edit Laporan Tersimpan",
       'Ada laporan berstatus "Tersimpan"',
       "1. Edit data realisasi\n2. Simpan",
       f"Baris: {top3_names[0][:35]}\nRealisasi Rp diubah dari {int(top3_pagu[0]*0.08):,.0f} menjadi {int(top3_pagu[0]*0.10):,.0f}",
       f'"Laporan berhasil diperbarui"\nRealisasi Rp: {fmt_rp(int(top3_pagu[0]*0.10))}')

    tc("Laporan", "Input Realisasi Fisik > 100%",
       "Di halaman input laporan",
       "1. Isi Realisasi Fisik > 100\n2. Simpan",
       "Realisasi Fisik: 150",
       'Validasi gagal\n"Realisasi fisik tidak boleh lebih dari 100%"')

    tc("Laporan", "Input Saat Izin Tidak Aktif",
       "Izin edit laporan TIDAK aktif",
       "1. Coba isi realisasi\n2. Simpan",
       f"Realisasi K: 10, Realisasi Rp: {int(top3_pagu[0]*0.15):,.0f}",
       '"Anda tidak diizinkan untuk mengedit laporan saat ini"\nForm read-only')

    # --- SUBMIT ---
    write_section_row(ws, r, "6. SUBMIT LAPORAN", len(TC_HEADERS)); r += 1

    tc("Submit", "Submit Laporan ke Admin",
       'Ada laporan Januari 2026 berstatus "Tersimpan"',
       '1. Pastikan semua baris terisi\n2. Klik "Kirim"\n3. Konfirmasi',
       'Klik "Kirim" untuk Januari 2026',
       '"Laporan berhasil dikirim"\nStatus: Tersimpan -> Terkirim\nLaporan read-only\nMuncul di sisi admin untuk verifikasi')

    tc("Submit", "Edit Laporan Terkirim",
       'Laporan berstatus "Terkirim"',
       "1. Coba edit realisasi",
       "Coba ubah Realisasi Rp",
       'Tidak bisa edit\nForm read-only\n"Laporan sudah dikirim, tidak dapat diedit"')

    tc("Submit", "Edit Laporan Dikembalikan",
       "Admin mengembalikan laporan dengan catatan",
       "1. Buka laporan yang dikembalikan\n2. Lihat catatan admin\n3. Perbaiki data\n4. Simpan",
       f'Catatan admin: "Data realisasi Rp tidak sesuai"\nPerbaikan: Realisasi Rp diubah',
       'Catatan admin tampil\nBisa edit kembali\n"Laporan berhasil diperbarui"\nBisa kirim ulang')

    # --- IZIN EDIT ---
    write_section_row(ws, r, "7. CEK IZIN EDIT", len(TC_HEADERS)); r += 1

    tc("Izin Edit", "Cek Izin Edit - Aktif",
       "Admin sudah buka izin edit",
       "1. Buka halaman Laporan",
       "(otomatis dicek)",
       "Form input aktif\nTombol Simpan dan Kirim tersedia")

    tc("Izin Edit", "Cek Izin Edit - Tidak Aktif",
       "Admin belum/sudah tutup izin edit",
       "1. Buka halaman Laporan",
       "(otomatis dicek)",
       'Form read-only\n"Pengisian laporan belum dibuka oleh admin"\nTombol disabled')

    # --- EXPORT ---
    write_section_row(ws, r, "8. EXPORT LAPORAN", len(TC_HEADERS)); r += 1

    tc("Export", "Export Laporan ke Excel",
       f"Login sebagai {nama}, ada data laporan",
       '1. Klik "Export"\n2. Pilih filter',
       "Tahun: 2026",
       f"File: laporan_{username}_2026.xlsx\nHeader: Laporan Puskesmas {nama} - 2026\nKolom: Sub Kegiatan, Sumber, Satuan, Target K, Target Rp, Realisasi K, Realisasi Rp, Realisasi Fisik, Permasalahan, Upaya")

    tc("Export", "Export per Bulan",
       "Ada data laporan Januari 2026",
       "1. Filter Tahun 2026, Bulan Januari\n2. Klik Export",
       "Tahun: 2026, Bulan: Januari",
       f"File hanya berisi data Januari 2026 milik {nama}")

    # --- CARA PENGISIAN ---
    write_section_row(ws, r, "9. CARA PENGISIAN", len(TC_HEADERS)); r += 1

    tc("Panduan", "Buka Halaman Cara Pengisian",
       f"Login sebagai {nama}",
       '1. Klik menu "Cara Pengisian"',
       "Klik sidebar > Cara Pengisian",
       "Halaman panduan tampil:\n- Cara login\n- Cara input target\n- Cara input angkas\n- Cara input laporan\n- Cara submit\n- Cara export")

    # --- LOGOUT ---
    write_section_row(ws, r, "10. LOGOUT", len(TC_HEADERS)); r += 1

    tc("Logout", f"Logout {nama}",
       f"Login sebagai {nama}",
       "1. Klik tombol Logout",
       "Klik Logout",
       "Token JWT dihapus\nRedirect ke /login")

    # ---- Sheet 2: Data Target (Real) ----
    ws2 = wb.create_sheet("Data Target (Real)")
    set_col_widths(ws2, [5, 22, 50, 25, 20])

    write_title(ws2, 1, f"DATA TARGET ANGGARAN - PUSKESMAS {nama.upper()} (TAHUN 2026)", 5)
    write_subtitle(ws2, 2, f"Total Pagu: {fmt_rp(total_pagu)} | Jumlah Sub Kegiatan x Sumber: {len(data)}", 5)

    write_header_row(ws2, 4, ["No", "Kode Sub Kegiatan", "Nama Sub Kegiatan", "Sumber Anggaran", "Pagu (Rp)"])

    for i, d in enumerate(data):
        row_data = [i + 1, d["kode"], d["nama"], d["sumber"], d["pagu"]]
        write_data_row(ws2, 5 + i, row_data)
        ws2.cell(row=5+i, column=5).number_format = '#,##0'

    # Total row
    total_row = 5 + len(data)
    ws2.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    cell = ws2.cell(row=total_row, column=1, value="TOTAL PAGU")
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws2.cell(row=total_row, column=c).border = THIN_BORDER
    cell_total = ws2.cell(row=total_row, column=5, value=total_pagu)
    cell_total.font = Font(bold=True)
    cell_total.border = THIN_BORDER
    cell_total.number_format = '#,##0'

    # ---- Sheet 3: Ringkasan ----
    ws3 = wb.create_sheet("Ringkasan")
    set_col_widths(ws3, [20, 15, 10])

    write_title(ws3, 1, "RINGKASAN TEST CASE", 3)
    write_header_row(ws3, 3, ["Modul", "Jumlah TC", "ID Range"])

    modules = [
        ("Login & Autentikasi", 4, f"{tc_prefix}-001 s/d {tc_prefix}-004"),
        ("Dashboard Puskesmas", 4, f"{tc_prefix}-005 s/d {tc_prefix}-008"),
        ("Manajemen Target", 6, f"{tc_prefix}-009 s/d {tc_prefix}-014"),
        ("Manajemen Angkas", 3, f"{tc_prefix}-015 s/d {tc_prefix}-017"),
        ("Input Laporan", 6, f"{tc_prefix}-018 s/d {tc_prefix}-023"),
        ("Submit Laporan", 3, f"{tc_prefix}-024 s/d {tc_prefix}-026"),
        ("Cek Izin Edit", 2, f"{tc_prefix}-027 s/d {tc_prefix}-028"),
        ("Export Laporan", 2, f"{tc_prefix}-029 s/d {tc_prefix}-030"),
        ("Cara Pengisian", 1, f"{tc_prefix}-031"),
        ("Logout", 1, f"{tc_prefix}-032"),
    ]
    for i, (mod, cnt, ids) in enumerate(modules):
        write_data_row(ws3, 4 + i, [mod, cnt, ids])

    total_r = 4 + len(modules)
    write_data_row(ws3, total_r, ["TOTAL", tc_num, ""])
    ws3.cell(row=total_r, column=1).font = Font(bold=True)
    ws3.cell(row=total_r, column=2).font = Font(bold=True)

    wb.save(filename)
    print(f"Generated: {filename} ({tc_num} test cases)")


# ============================================================
# ADMIN TEST DOC GENERATOR
# ============================================================

def generate_admin_doc(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    set_col_widths(ws, TC_WIDTHS)

    r = 1
    write_title(ws, r, "DOKUMEN PENGUJIAN - ADMIN (DINKES)", len(TC_HEADERS)); r += 1
    write_subtitle(ws, r, "Aplikasi: E-EVKIN Modern | Role: Admin (Dinas Kesehatan Kab. Bogor) | Username: dinkes | Password: dinkes", len(TC_HEADERS)); r += 1
    write_subtitle(ws, r, "Tanggal: 30 Maret 2026 | Tahun Anggaran: 2026 | Total Puskesmas: 106", len(TC_HEADERS)); r += 1
    r += 1

    write_header_row(ws, r, TC_HEADERS); r += 1

    tc_num = 0
    def tc(modul, nama_tc, prasyarat, langkah, contoh, output):
        nonlocal r, tc_num
        tc_num += 1
        write_data_row(ws, r, [tc_num, f"TC-ADM-{tc_num:03d}", modul, nama_tc, prasyarat, langkah, contoh, output, "", ""])
        r += 1

    # --- LOGIN ---
    write_section_row(ws, r, "1. LOGIN & AUTENTIKASI", len(TC_HEADERS)); r += 1

    tc("Login", "Login Admin Berhasil",
       "Aplikasi dapat diakses, belum login",
       "1. Buka /login\n2. Input username: dinkes\n3. Input password: dinkes\n4. Klik Login",
       "Username: dinkes\nPassword: dinkes",
       'Redirect ke /dashboard\nSidebar: Dashboard, Master Data, Puskesmas, Konfigurasi, Laporan, Target Upload, Target Edit\nHeader: "Dinas Kesehatan"')

    tc("Login", "Login Gagal - Password Salah",
       "Belum login",
       "1. Input username: dinkes\n2. Input password: salah123\n3. Klik Login",
       "Username: dinkes\nPassword: salah123",
       '"Username atau password salah"')

    tc("Login", "Login Gagal - Field Kosong",
       "Belum login",
       "1. Biarkan kosong\n2. Klik Login",
       "Username: (kosong)\nPassword: (kosong)",
       '"Username wajib diisi"\n"Password wajib diisi"')

    tc("Login", "Verifikasi Token (Refresh)",
       "Sudah login sebagai admin",
       "1. Tekan F5",
       "(otomatis)",
       "Tetap login\nDashboard admin tampil kembali")

    tc("Login", "Akses Tanpa Login",
       "Belum login / token expired",
       "1. Buka /dashboard langsung",
       "URL: /dashboard",
       'Redirect ke /login\n"Silakan login terlebih dahulu"')

    # --- DASHBOARD ---
    write_section_row(ws, r, "2. DASHBOARD ADMIN", len(TC_HEADERS)); r += 1

    tc("Dashboard", "Tampilan Awal Dashboard",
       "Login sebagai admin",
       '1. Klik menu "Dashboard"',
       "Klik sidebar > Dashboard",
       f"Filter: Tahun, Bulan\nKartu Statistik: Total Laporan, Terkirim, Pending, Terverifikasi\nGrafik: Target vs Realisasi Anggaran\nTop/Bottom 10 penyerapan\nStatus Pelaporan per puskesmas\n\nContoh referensi data:\n- Bojonggede total pagu: {fmt_rp(TOTAL_BJG)}\n- Cibinong total pagu: {fmt_rp(TOTAL_CBN)}")

    tc("Dashboard", "Filter Dashboard per Tahun & Bulan",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026\n2. Pilih Bulan: Januari",
       "Tahun: 2026, Bulan: Januari",
       "Data di-refresh sesuai filter\nTop 10 & Bottom 10 berubah")

    tc("Dashboard", "Grafik Budget YTD",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026\n2. Lihat grafik YTD",
       "Tahun: 2026",
       "Grafik bar + line (dual Y-axis)\nY-kiri: Target Anggaran (Rp)\nY-kanan: Realisasi Fisik (%)\n100% sejajar target tertinggi")

    tc("Dashboard", "Filter Sumber Anggaran",
       "Di halaman Dashboard",
       "1. Pilih Tahun: 2026\n2. Filter Sumber: BLUD Puskesmas",
       "Tahun: 2026, Sumber: BLUD Puskesmas",
       "Grafik hanya data BLUD\nNilai target/realisasi subset dari total")

    tc("Dashboard", "Status Pelaporan Puskesmas",
       "Di Dashboard, Tahun 2026, Bulan Januari",
       "1. Scroll ke status pelaporan",
       "(otomatis)",
       "Sudah Mengirim: daftar puskesmas\nBelum Mengirim: daftar puskesmas\nTotal: 106 puskesmas")

    # --- MANAJEMEN PUSKESMAS ---
    write_section_row(ws, r, "3. MANAJEMEN PENGGUNA PUSKESMAS", len(TC_HEADERS)); r += 1

    tc("Puskesmas", "Lihat Daftar Puskesmas",
       "Login sebagai admin",
       '1. Klik menu "Puskesmas"',
       "Klik sidebar > Puskesmas",
       "Tabel: No, Nama, Username, Kecamatan, Wilayah, ID BLUD, Aksi\nContoh:\n- Bojonggede | bojonggede | Bojonggede | Parung | BLUD\n- Cibinong | cibinong | Cibinong | Cibinong | JKN\nTotal: 106 puskesmas")

    tc("Puskesmas", "Tambah Puskesmas",
       "Di halaman Puskesmas",
       '1. Klik "Tambah Puskesmas"\n2. Isi form\n3. Simpan',
       "Nama: Puskesmas Test\nUsername: testbaru\nPassword: test123\nKecamatan: Dramaga\nWilayah: Cibinong\nID BLUD: JKN",
       '"Puskesmas berhasil ditambahkan"\nTotal bertambah')

    tc("Puskesmas", "Edit Puskesmas",
       "Ada data puskesmas test",
       '1. Klik [Edit]\n2. Ubah kecamatan\n3. Simpan',
       "Kecamatan: Dramaga -> Bogor Barat",
       '"Data puskesmas berhasil diperbarui"')

    tc("Puskesmas", "Hapus Puskesmas",
       "Ada data puskesmas test",
       '1. Klik [Hapus]\n2. Konfirmasi OK',
       "Konfirmasi: OK",
       '"Puskesmas berhasil dihapus"\nBaris hilang dari tabel')

    tc("Puskesmas", "Username Duplikat",
       "Di halaman Puskesmas",
       '1. Tambah dengan username existing\n2. Simpan',
       "Username: cibinong (sudah ada)",
       '"Username sudah digunakan"\nData tidak masuk')

    # --- MASTER DATA SATUAN ---
    write_section_row(ws, r, "4. MASTER DATA - SATUAN", len(TC_HEADERS)); r += 1

    tc("Satuan", "Lihat Daftar Satuan",
       "Login sebagai admin",
       '1. Klik "Master Data" > tab "Satuan"',
       "Klik tab Satuan",
       "Tabel: Orang, Dokumen, unit kerja, Laporan, Kegiatan")

    tc("Satuan", "Tambah Satuan",
       "Di Master Data > Satuan",
       '1. Klik "Tambah Satuan"\n2. Isi nama\n3. Simpan',
       "Nama: Paket",
       '"Satuan berhasil ditambahkan"\nBaris baru: Paket')

    tc("Satuan", "Edit Satuan",
       "Satuan 'Paket' sudah ada",
       '1. Klik [Edit]\n2. Ubah nama\n3. Simpan',
       "Paket -> Paket Kegiatan",
       '"Satuan berhasil diperbarui"')

    tc("Satuan", "Hapus Satuan",
       "Satuan tidak dipakai di laporan",
       '1. Klik [Hapus]\n2. Konfirmasi',
       "Konfirmasi: OK",
       '"Satuan berhasil dihapus"')

    # --- MASTER DATA SUMBER ANGGARAN ---
    write_section_row(ws, r, "5. MASTER DATA - SUMBER ANGGARAN", len(TC_HEADERS)); r += 1

    tc("Sumber Anggaran", "Lihat Daftar Sumber Anggaran",
       "Login sebagai admin",
       '1. Klik "Master Data" > tab "Sumber Anggaran"',
       "Klik tab Sumber Anggaran",
       "Tabel:\n1 | BLUD Puskesmas\n2 | DAK Non Fisik\n3 | APBD\n4 | JKN (Dana Kapitasi)")

    tc("Sumber Anggaran", "Tambah Sumber Anggaran",
       "Di Master Data > Sumber Anggaran",
       '1. Klik "Tambah"\n2. Isi nama\n3. Simpan',
       "Nama: BOK Tambahan",
       '"Sumber anggaran berhasil ditambahkan"')

    tc("Sumber Anggaran", "Edit Sumber Anggaran",
       "'BOK Tambahan' sudah ada",
       '1. Klik [Edit]\n2. Ubah nama\n3. Simpan',
       "BOK Tambahan -> BOK Puskesmas",
       '"Sumber anggaran berhasil diperbarui"')

    tc("Sumber Anggaran", "Hapus Sumber Anggaran",
       "Tidak dipakai di laporan",
       '1. Klik [Hapus]\n2. Konfirmasi',
       "Konfirmasi: OK",
       '"Sumber anggaran berhasil dihapus"')

    # --- MASTER DATA KEGIATAN ---
    write_section_row(ws, r, "6. MASTER DATA - KEGIATAN", len(TC_HEADERS)); r += 1

    tc("Kegiatan", "Lihat Daftar Kegiatan",
       "Login sebagai admin",
       '1. Klik "Master Data" > tab "Kegiatan"',
       "Klik tab Kegiatan",
       "Tabel:\n- 1.02.01.2.10 | Peningkatan Pelayanan BLUD\n- 1.02.02.2.02 | Penyediaan Layanan Kesehatan untuk UKM dan UKP\n- 1.02.05.2.03 | Pengembangan dan Pelaksanaan UKBM\n- 1.02.05.2.01 | Penguatan Pelayanan Kesehatan Primer")

    tc("Kegiatan", "Tambah Kegiatan",
       "Di Master Data > Kegiatan",
       '1. Klik "Tambah"\n2. Isi form\n3. Simpan',
       "Kode: 1.02.06.2.01\nNama: Kegiatan Pengujian Sistem",
       '"Kegiatan berhasil ditambahkan"')

    tc("Kegiatan", "Hapus Kegiatan Berisi Sub",
       "Kegiatan punya sub kegiatan",
       '1. Klik [Hapus] pada 1.02.02.2.02\n2. Konfirmasi',
       "Konfirmasi: OK",
       '"Tidak dapat menghapus kegiatan yang masih memiliki sub kegiatan"')

    tc("Kegiatan", "Hapus Kegiatan Kosong",
       "Kegiatan tidak punya sub",
       '1. Klik [Hapus] pada Kegiatan Pengujian\n2. Konfirmasi',
       "Konfirmasi: OK",
       '"Kegiatan berhasil dihapus"')

    # --- MASTER DATA SUB KEGIATAN ---
    write_section_row(ws, r, "7. MASTER DATA - SUB KEGIATAN", len(TC_HEADERS)); r += 1

    tc("Sub Kegiatan", "Lihat Daftar Sub Kegiatan",
       "Login sebagai admin",
       '1. Klik "Master Data" > tab "Sub Kegiatan"',
       "Klik tab Sub Kegiatan",
       "Tabel contoh:\n- 1.02.01.2.10.0001 | Pelayanan dan Penunjang Pelayanan BLUD\n- 1.02.02.2.02.0001 | Pengelolaan Pelayanan Kesehatan Ibu Hamil\n- 1.02.02.2.02.0033 | Operasional Pelayanan Puskesmas")

    tc("Sub Kegiatan", "Tambah Sub Kegiatan",
       "Di Master Data > Sub Kegiatan",
       '1. Klik "Tambah"\n2. Isi form\n3. Simpan',
       "Kegiatan Induk: Peningkatan Pelayanan BLUD\nKode: 1.02.01.2.10.0099\nNama: Sub Kegiatan Test\nIndikator: Jumlah dokumen pengujian",
       '"Sub kegiatan berhasil ditambahkan"')

    tc("Sub Kegiatan", "Edit Sub Kegiatan",
       "Sub Kegiatan Test sudah ada",
       '1. Klik [Edit]\n2. Ubah indikator\n3. Simpan',
       "Indikator: Jumlah laporan pengujian yang divalidasi",
       '"Sub kegiatan berhasil diperbarui"')

    tc("Sub Kegiatan", "Hapus Sub Kegiatan",
       "Sub Kegiatan Test tidak terkait laporan",
       '1. Klik [Hapus]\n2. Konfirmasi',
       "Konfirmasi: OK",
       '"Sub kegiatan berhasil dihapus"')

    # --- KONFIGURASI ASSIGN ---
    write_section_row(ws, r, "8. KONFIGURASI - ASSIGN SUB KEGIATAN", len(TC_HEADERS)); r += 1

    tc("Assign", "Lihat Overview Assignment",
       "Login sebagai admin",
       '1. Klik menu "Konfigurasi"',
       "Klik sidebar > Konfigurasi",
       "Tabel overview:\n- Bojonggede | 16 sub kegiatan assigned\n- Cibinong | 16 sub kegiatan assigned\n- ... (semua 106 puskesmas)")

    tc("Assign", "Assign Sub Kegiatan",
       "Di halaman Konfigurasi",
       '1. Klik "Cibinong"\n2. Centang sub kegiatan\n3. Simpan',
       "Centang:\n[x] Pelayanan BLUD\n[x] Pelayanan Ibu Hamil\n[x] Operasional Puskesmas",
       '"Sub kegiatan berhasil di-assign"')

    tc("Assign", "Hapus Assignment",
       "Cibinong sudah punya assignment",
       '1. Klik "Cibinong"\n2. Hapus centang\n3. Simpan',
       "Hapus centang: Pelayanan BLUD",
       "Assignment berkurang\nSub kegiatan tidak muncul saat input laporan")

    # --- IZIN EDIT ---
    write_section_row(ws, r, "9. KONFIGURASI - IZIN EDIT", len(TC_HEADERS)); r += 1

    tc("Izin Edit", "Buka Izin Edit Laporan (Global)",
       "Login sebagai admin",
       "1. Buka pengaturan izin edit\n2. Isi form\n3. Simpan",
       "Scope: laporan\nTahun: 2026\nBulan: Januari\nUser: (semua)\nEnabled: Ya",
       '"Izin edit berhasil disimpan"\nSemua puskesmas bisa input Januari 2026')

    tc("Izin Edit", "Buka Izin Edit dengan Time Window",
       "Login sebagai admin",
       "1. Isi form dengan waktu\n2. Simpan",
       "Scope: laporan\nTahun: 2026, Bulan: Februari\nStart: 2026-03-01 08:00\nEnd: 2026-03-15 17:00",
       "Izin hanya aktif 1-15 Maret 2026\nDi luar waktu: form read-only")

    tc("Izin Edit", "Buka Izin Edit Target Kinerja",
       "Login sebagai admin",
       "1. Isi form\n2. Simpan",
       "Scope: target_kinerja\nTahun: 2026\nEnabled: Ya",
       "Semua puskesmas bisa edit target kinerja 2026")

    tc("Izin Edit", "Buka Izin Edit Target Rp",
       "Login sebagai admin",
       "1. Isi form\n2. Simpan",
       "Scope: target_rp\nTahun: 2026\nEnabled: Ya",
       "Semua puskesmas bisa edit target Rp 2026")

    tc("Izin Edit", "Buka Izin Edit Angkas",
       "Login sebagai admin",
       "1. Isi form\n2. Simpan",
       "Scope: angkas\nTahun: 2026\nEnabled: Ya",
       "Semua puskesmas bisa edit angkas 2026")

    tc("Izin Edit", "Tutup Izin Edit Laporan",
       "Izin edit sedang aktif",
       "1. Ubah enabled\n2. Simpan",
       "Scope: laporan\nTahun: 2026, Bulan: Januari\nEnabled: Tidak",
       '"Izin edit berhasil diperbarui"\nPuskesmas tidak bisa input')

    tc("Izin Edit", "Izin Edit per Puskesmas",
       "Login sebagai admin",
       "1. Isi form per puskesmas\n2. Simpan",
       "Scope: laporan\nUser: Cibinong\nTahun: 2026, Bulan: Maret\nEnabled: Ya",
       "Hanya Cibinong bisa edit Maret 2026\nBojonggede tidak bisa")

    # --- VERIFIKASI ---
    write_section_row(ws, r, "10. VERIFIKASI LAPORAN", len(TC_HEADERS)); r += 1

    tc("Verifikasi", "Lihat Laporan Terkirim",
       "Login sebagai admin, ada laporan terkirim",
       '1. Klik menu "Laporan"',
       "Klik sidebar > Laporan",
       "Tabel grouped per puskesmas:\n- Cibinong | Jan 2026 | Terkirim | 16 sub kegiatan\n- Bojonggede | Jan 2026 | Terkirim | 16 sub kegiatan")

    tc("Verifikasi", "Filter Laporan",
       "Di halaman Laporan admin",
       "1. Set filter\n2. Cari",
       "Puskesmas: Cibinong\nTahun: 2026, Bulan: Januari",
       "Hanya laporan Cibinong Januari 2026")

    tc("Verifikasi", "Lihat Detail Laporan",
       "Di halaman Laporan",
       "1. Klik baris Cibinong Januari 2026",
       "Klik baris Cibinong | Jan 2026",
       f"Detail tabel:\nSub Kegiatan: Pelayanan BLUD\nSumber: BLUD Puskesmas\nTarget Rp: {fmt_rp(2_135_455_000)}\nRealisasi Rp, Fisik %, Permasalahan, Upaya")

    tc("Verifikasi", "Kembalikan Laporan",
       'Laporan berstatus "terkirim"',
       '1. Buka detail\n2. Klik "Kembalikan"\n3. Isi catatan\n4. Konfirmasi',
       'Catatan: "Data realisasi Rp tidak sesuai bukti"',
       'Status: Terkirim -> Tersimpan\nCatatan tampil di sisi puskesmas\n"Laporan berhasil dikembalikan"')

    tc("Verifikasi", "Bulk Return Laporan",
       "Ada beberapa laporan terkirim",
       '1. Pilih puskesmas/periode\n2. Klik "Kembalikan Semua"\n3. Konfirmasi',
       "Puskesmas: Bojonggede\nBulan: Januari 2026",
       '"16 laporan berhasil dikembalikan"\nSemua status -> Tersimpan')

    # --- TARGET ADMIN ---
    write_section_row(ws, r, "11. MANAJEMEN TARGET (ADMIN)", len(TC_HEADERS)); r += 1

    tc("Target Admin", "Lihat Semua Target",
       "Login sebagai admin",
       '1. Klik menu "Target Edit"',
       "Klik sidebar > Target Edit",
       f"Tabel target semua puskesmas:\n- Bojonggede | Pelayanan BLUD | BLUD | Target Rp: {fmt_rp(3_738_528_000)}\n- Cibinong | Pelayanan BLUD | BLUD | Target Rp: {fmt_rp(2_135_455_000)}\n- ...")

    tc("Target Admin", "Edit Target Puskesmas",
       "Di halaman Target Edit",
       '1. Klik [Edit] target Bojonggede\n2. Ubah target K\n3. Simpan',
       "Target K: 12 -> 15",
       '"Target berhasil diperbarui"\nRiwayat tercatat')

    tc("Target Admin", "Buat Target Baru",
       "Di halaman Target Edit",
       '1. Klik "Tambah Target"\n2. Isi form\n3. Simpan',
       f"Puskesmas: Cibinong\nSub: Operasional Puskesmas\nSumber: DAK Non Fisik (BOK)\nTarget K: 12\nTarget Rp: 77.276.000\nTahun: 2026",
       '"Target berhasil dibuat"')

    tc("Target Admin", "Lihat Riwayat Target",
       "Target pernah diubah",
       "1. Buka riwayat target",
       "Filter: Bojonggede",
       "Riwayat perubahan:\n- Versi 2: Target K 15, oleh dinkes\n- Versi 1: Target K 12, oleh bojonggede")

    # --- UPLOAD EXCEL ---
    write_section_row(ws, r, "12. UPLOAD TARGET DARI EXCEL", len(TC_HEADERS)); r += 1

    tc("Upload", "Upload Excel Berhasil",
       "Login sebagai admin, file Excel sesuai format",
       '1. Klik "Target Upload"\n2. Pilih file\n3. Tunggu proses',
       "File: Rekap_Ver3.xlsx\n(berisi data target semua puskesmas)",
       "Progress bar\nHasil: Berhasil X target, Gagal Y target\nDetail error per baris")

    tc("Upload", "Upload File Non-Excel",
       "Login sebagai admin",
       "1. Upload file PDF",
       "File: dokumen.pdf",
       '"Format file tidak didukung. Gunakan file Excel (.xlsx)"')

    # --- EXPORT ---
    write_section_row(ws, r, "13. EXPORT LAPORAN", len(TC_HEADERS)); r += 1

    tc("Export", "Export Semua Laporan",
       "Login sebagai admin, ada data",
       '1. Klik "Export"\n2. Pilih tahun',
       "Tahun: 2026",
       "File: laporan_2026.xlsx\nSheet 1: Summary semua puskesmas\nSheet 2+: Per puskesmas\nFormat: Rp, %, border, header")

    tc("Export", "Export dengan Filter",
       "Login sebagai admin",
       "1. Set filter\n2. Export",
       "Puskesmas: Cibinong\nTahun: 2026, Bulan: Januari",
       "File hanya berisi Cibinong Januari 2026")

    # --- AI CHAT ---
    write_section_row(ws, r, "14. AI CHAT ASSISTANT", len(TC_HEADERS)); r += 1

    tc("AI Chat", "Buka Chat Widget",
       "Login sebagai admin",
       "1. Klik ikon chat di pojok kanan bawah",
       "Klik ikon chat",
       "Window chat terbuka\nSuggested questions tampil")

    tc("AI Chat", "Tanya AI",
       "Chat widget terbuka",
       "1. Ketik pertanyaan\n2. Klik kirim",
       '"Puskesmas mana yang penyerapan anggarannya paling rendah bulan Januari 2026?"',
       "AI menjawab berdasarkan data:\n- Daftar puskesmas dengan penyerapan terendah\n- Menyebutkan angka Rp dan %")

    # --- REPORT ---
    write_section_row(ws, r, "15. LAPORAN AGREGAT", len(TC_HEADERS)); r += 1

    tc("Report", "Report by Sub Kegiatan",
       "Login sebagai admin, ada data",
       "1. Akses report by sub kegiatan\n2. Set filter",
       "Tahun: 2026, Bulan: Januari\nSub: Operasional Pelayanan Puskesmas",
       "Tabel agregat:\nTotal Target Rp dari semua puskesmas\nTotal Realisasi Rp\n% Penyerapan")

    tc("Report", "Report by Sumber Anggaran",
       "Login sebagai admin, ada data",
       "1. Akses report by sumber\n2. Set filter",
       "Tahun: 2026, Bulan: Januari\nSumber: BLUD Puskesmas",
       "Tabel agregat:\nTotal Target Rp BLUD\nTotal Realisasi Rp\n% Penyerapan")

    # --- HEALTH CHECK ---
    write_section_row(ws, r, "16. HEALTH CHECK", len(TC_HEADERS)); r += 1

    tc("Health Check", "Health Check Basic",
       "Server backend berjalan",
       "1. GET /health",
       "GET /health",
       '{"status":"ok","timestamp":"..."}\nHTTP 200')

    tc("Health Check", "Health Check Detail",
       "Akses dari localhost",
       "1. GET /health?detail=true",
       "GET /health?detail=true",
       '{"status":"ok","uptime":...,"memory":{...},"dbPool":{...},"cache":{...}}\nHTTP 200')

    # --- LOGOUT ---
    write_section_row(ws, r, "17. LOGOUT", len(TC_HEADERS)); r += 1

    tc("Logout", "Logout Admin",
       "Login sebagai admin",
       "1. Klik tombol Logout",
       "Klik Logout",
       "Token dihapus\nRedirect ke /login")

    # ---- Sheet 2: Data Referensi ----
    ws2 = wb.create_sheet("Data Referensi")
    set_col_widths(ws2, [5, 22, 55, 25, 20])

    write_title(ws2, 1, "DATA TARGET ANGGARAN REFERENSI - PUSKESMAS CONTOH", 5)

    # Bojonggede section
    write_subtitle(ws2, 3, f"PUSKESMAS BOJONGGEDE - Total Pagu: {fmt_rp(TOTAL_BJG)}", 5)
    write_header_row(ws2, 4, ["No", "Kode Sub Kegiatan", "Nama Sub Kegiatan", "Sumber Anggaran", "Pagu (Rp)"])
    for i, d in enumerate(BOJONGGEDE_DATA):
        write_data_row(ws2, 5 + i, [i+1, d["kode"], d["nama"], d["sumber"], d["pagu"]])
        ws2.cell(row=5+i, column=5).number_format = '#,##0'

    br = 5 + len(BOJONGGEDE_DATA) + 2

    # Cibinong section
    write_subtitle(ws2, br, f"PUSKESMAS CIBINONG - Total Pagu: {fmt_rp(TOTAL_CBN)}", 5)
    write_header_row(ws2, br + 1, ["No", "Kode Sub Kegiatan", "Nama Sub Kegiatan", "Sumber Anggaran", "Pagu (Rp)"])
    for i, d in enumerate(CIBINONG_DATA):
        write_data_row(ws2, br + 2 + i, [i+1, d["kode"], d["nama"], d["sumber"], d["pagu"]])
        ws2.cell(row=br+2+i, column=5).number_format = '#,##0'

    # ---- Sheet 3: Ringkasan ----
    ws3 = wb.create_sheet("Ringkasan")
    set_col_widths(ws3, [30, 12, 20])

    write_title(ws3, 1, "RINGKASAN TEST CASE ADMIN", 3)
    write_header_row(ws3, 3, ["Modul", "Jumlah TC", "ID Range"])

    modules = [
        ("Login & Autentikasi", 5),
        ("Dashboard Admin", 5),
        ("Manajemen Puskesmas", 5),
        ("Master Data - Satuan", 4),
        ("Master Data - Sumber Anggaran", 4),
        ("Master Data - Kegiatan", 4),
        ("Master Data - Sub Kegiatan", 4),
        ("Konfigurasi - Assign", 3),
        ("Konfigurasi - Izin Edit", 7),
        ("Verifikasi Laporan", 5),
        ("Manajemen Target", 4),
        ("Upload Target Excel", 2),
        ("Export Laporan", 2),
        ("AI Chat", 2),
        ("Laporan Agregat", 2),
        ("Health Check", 2),
        ("Logout", 1),
    ]

    start = 1
    for i, (mod, cnt) in enumerate(modules):
        end = start + cnt - 1
        ids = f"TC-ADM-{start:03d} s/d TC-ADM-{end:03d}" if cnt > 1 else f"TC-ADM-{start:03d}"
        write_data_row(ws3, 4 + i, [mod, cnt, ids])
        start = end + 1

    total_r = 4 + len(modules)
    write_data_row(ws3, total_r, ["TOTAL", tc_num, ""])
    ws3.cell(row=total_r, column=1).font = Font(bold=True)
    ws3.cell(row=total_r, column=2).font = Font(bold=True)

    wb.save(filename)
    print(f"Generated: {filename} ({tc_num} test cases)")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    import os
    outdir = os.path.dirname(os.path.abspath(__file__))

    generate_puskesmas_doc(
        os.path.join(outdir, "TEST_PUSKESMAS_BOJONGGEDE.xlsx"),
        nama="Bojonggede", username="bojonggede",
        kecamatan="Bojonggede", wilayah="Parung",
        data=BOJONGGEDE_DATA, total_pagu=TOTAL_BJG,
        tc_prefix="TC-BJG"
    )

    generate_puskesmas_doc(
        os.path.join(outdir, "TEST_PUSKESMAS_CIBINONG.xlsx"),
        nama="Cibinong", username="cibinong",
        kecamatan="Cibinong", wilayah="Cibinong",
        data=CIBINONG_DATA, total_pagu=TOTAL_CBN,
        tc_prefix="TC-CBN"
    )

    generate_admin_doc(os.path.join(outdir, "TEST_ADMIN_DINKES.xlsx"))

    print("\nDone! All 3 Excel test documents generated.")
