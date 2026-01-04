import openpyxl
import json

# Data dari Excel 
wb = openpyxl.load_workbook('d:/proj/e-evkin-modern/docs/Rekap_Ver3.xlsx')
sheet = wb.active

excel_data = {}
for row in range(2, sheet.max_row + 1):
    kode = sheet.cell(row=row, column=7).value
    nama = sheet.cell(row=row, column=8).value
    if kode and nama:
        if 'uskesmas' in str(nama).lower() or 'labkesda' in str(nama).lower():
            # Extract puskesmas name (remove 'Puskesmas ' prefix)
            nama_clean = str(nama).strip()
            if nama_clean.lower().startswith('puskesmas '):
                nama_short = nama_clean[10:].strip()
            elif nama_clean.lower().startswith('puskemas '):  # typo
                nama_short = nama_clean[9:].strip()
            else:
                nama_short = nama_clean
            excel_data[kode] = {'nama_full': nama_clean, 'nama_short': nama_short}

# Data puskesmas dari seeder - extracted manually
seeder_puskesmas = [
    'Bojonggede', 'Bagoang', 'Jasinga', 'Curug', 'Cigudeg', 'Lebakwangi', 'Bunar',
    'Sukajaya', 'Kiara Pandak', 'Parung Panjang', 'Dago', 'Tenjo', 'Pasar Rebo',
    'Nanggung', 'Curug Bitung', 'Leuwiliang', 'Puraseda', 'Leuwisadeng', 'Sadeng Pasar',
    'Rumpin', 'Gobang', 'Cicangkal', 'Cibungbulang', 'Cijujung', 'Situ Udik',
    'Pamijahan', 'Ciasmara', 'Cibening', 'Ciampea', 'Ciampea Udik', 'Pasir',
    'Cihideung Udik', 'Tenjolaya', 'Ciomas', 'Ciapus', 'Laladon', 'Kota Batu',
    'Dramaga', 'Kampung Manggis', 'Purwasari', 'Cangkurawok', 'Sirnagalih', 'Tamansari',
    'Sukaresmi', 'Ciawi', 'Banjarsari', 'Citapen', 'Cisarua', 'Cibulan', 'Megamendung',
    'Sukamanah', 'Caringin', 'Ciderum', 'Cinagara', 'Cijeruk', 'Sukaharja', 'Cigombong',
    'Ciburayut', 'Parung', 'Cogreg', 'Gunung Sindur', 'Suliwer', 'Kemang', 'Jampang',
    'Tajurhalang', 'Kemuning', 'Ragajaya', 'Ciseeng', 'Cibeuteung Udik', 'Bantarjaya',
    'Rancabungur', 'Cirimekar', 'Cibinong', 'Karadenan', 'Pabuaran Indah', 'Gunung Putri',
    'Ciangsana', 'Karanggan', 'Bojong Nangka', 'Cimandala', 'Sukaraja', 'Cilebut',
    'Babakan Madang', 'Sentul', 'Cijayanti', 'Citeureup', 'Leuwinutug', 'Tajur',
    'Jonggol', 'Sukanegara', 'Balekambang', 'Cileungsi', 'Pasir Angin', 'Gandoang',
    'Sukamakmur', 'Sukadamai', 'Klapanunggal', 'Bojong', 'Cariu', 'Karyamekar',
    'Tanjungsari', 'Labkesda'
]

# Normalize function
def normalize(s):
    return s.lower().replace(' ', '').replace('-', '').replace('_', '')

# Create mapping
seeder_normalized = {normalize(n): n for n in seeder_puskesmas}

print('=' * 80)
print('ANALISIS MAPPING KODE SUB UNIT')
print('=' * 80)
print(f'Total entries dari Excel/PDF: {len(excel_data)}')
print(f'Total Puskesmas di seeder: {len(seeder_puskesmas)}')
print()

matched = []
not_matched_excel = []

for kode, data in sorted(excel_data.items()):
    nama_short = data['nama_short']
    nama_norm = normalize(nama_short)
    
    if nama_norm in seeder_normalized:
        matched.append({
            'kode_sub_unit': kode,
            'nama_excel': data['nama_full'],
            'nama_seeder': seeder_normalized[nama_norm]
        })
    else:
        # Try partial match
        found = False
        for norm_key, orig_name in seeder_normalized.items():
            if nama_norm in norm_key or norm_key in nama_norm:
                matched.append({
                    'kode_sub_unit': kode,
                    'nama_excel': data['nama_full'],
                    'nama_seeder': orig_name,
                    'note': 'partial match'
                })
                found = True
                break
        if not found:
            not_matched_excel.append({
                'kode_sub_unit': kode,
                'nama_excel': data['nama_full'],
                'nama_short': nama_short
            })

# Find seeder puskesmas not in Excel
seeder_matched = set(normalize(m['nama_seeder']) for m in matched)
not_in_excel = []
for p in seeder_puskesmas:
    if normalize(p) not in seeder_matched:
        not_in_excel.append(p)

print('=' * 60)
print('PUSKESMAS YANG BERHASIL DICOCOKKAN')
print('=' * 60)
for m in sorted(matched, key=lambda x: x['kode_sub_unit']):
    note = f" ({m.get('note', '')})" if m.get('note') else ''
    print(f"{m['kode_sub_unit']} => {m['nama_seeder']}{note}")

print()
print('=' * 60)
print('PUSKESMAS DI EXCEL YANG TIDAK DITEMUKAN DI SEEDER')
print('=' * 60)
for item in not_matched_excel:
    print(f"{item['kode_sub_unit']} => {item['nama_excel']}")

print()
print('=' * 60)
print('PUSKESMAS DI SEEDER YANG TIDAK ADA DI EXCEL')
print('=' * 60)
for p in not_in_excel:
    print(f"  - {p}")

print()
print('=' * 60)
print('RINGKASAN')
print('=' * 60)
print(f'Berhasil dicocokkan: {len(matched)}')
print(f'Di Excel tapi tidak di seeder: {len(not_matched_excel)}')
print(f'Di seeder tapi tidak di Excel: {len(not_in_excel)}')

# Export untuk update database
print()
print('=' * 60)
print('MAPPING UNTUK UPDATE DATABASE (JSON)')
print('=' * 60)
mapping = {}
for m in matched:
    mapping[m['nama_seeder']] = m['kode_sub_unit']
print(json.dumps(mapping, indent=2))
